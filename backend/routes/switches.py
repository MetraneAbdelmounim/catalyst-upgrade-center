from flask import Blueprint, request, jsonify
from bson import ObjectId
from datetime import datetime, timezone
from models import switch_schema

switches_bp = Blueprint("switches", __name__, url_prefix="/api/switches")

def init_switches(db):
    @switches_bp.route("", methods=["GET"])
    def list_all():
        q = {}
        if request.args.get("site"):     q["site"] = request.args["site"]
        if request.args.get("status"):   q["status"] = request.args["status"]
        if request.args.get("platform"): q["platform"] = request.args["platform"]
        if request.args.get("search"):
            s = request.args["search"]
            q["$or"] = [{"name": {"$regex": s, "$options": "i"}},
                        {"ip_address": {"$regex": s, "$options": "i"}},
                        {"model": {"$regex": s, "$options": "i"}}]
        rows = list(db.switches.find(q).sort("name", 1))
        for r in rows: r["_id"] = str(r["_id"])
        return jsonify(rows)

    @switches_bp.route("", methods=["POST"])
    def create():
        d = request.json
        if not d.get("name") or not d.get("ip_address"):
            return jsonify({"error": "name and ip_address required"}), 400
        if db.switches.find_one({"ip_address": d["ip_address"]}):
            return jsonify({"error": "IP already exists"}), 409
        doc = switch_schema(d)
        res = db.switches.insert_one(doc)
        doc["_id"] = str(res.inserted_id)
        return jsonify(doc), 201

    # ── Static paths MUST come before /<sid> ─────────────────

    @switches_bp.route("/check-all", methods=["POST"])
    def check_all():
        """Ping all switches and update their online/offline status."""
        from services.health_checker import check_all_switches
        timeout = int(request.args.get("timeout", 2))
        results = check_all_switches(db, ping_timeout=timeout)
        online = sum(1 for r in results if r.get("new_status") == "online")
        offline = sum(1 for r in results if r.get("new_status") == "offline")
        changed = sum(1 for r in results if r.get("old_status") != r.get("new_status") and not r.get("skipped"))
        return jsonify({
            "total": len(results),
            "online": online,
            "offline": offline,
            "changed": changed,
            "results": results,
        })

    @switches_bp.route("/bulk-delete", methods=["POST"])
    def bulk_delete():
        """Delete multiple switches by ID list."""
        ids = request.json.get("ids", [])
        if not ids:
            return jsonify({"error": "No IDs provided"}), 400
        object_ids = [ObjectId(i) for i in ids]
        result = db.switches.delete_many({"_id": {"$in": object_ids}})
        return jsonify({"deleted": result.deleted_count})

    # ── Discovery job tracker (in-memory) ──────────────
    import threading as _threading
    _discovery_jobs = {}
    _discovery_lock = _threading.Lock()

    @switches_bp.route("/bulk-import", methods=["POST"])
    def bulk_import():
        """Import switches from XLSX + auto-discover in background with tracking."""
        import io, uuid
        from openpyxl import load_workbook

        if "file" not in request.files:
            return jsonify({"error": "No file uploaded"}), 400
        file = request.files["file"]
        if not file.filename.endswith((".xlsx", ".xls")):
            return jsonify({"error": "File must be .xlsx"}), 400

        try:
            wb = load_workbook(io.BytesIO(file.read()), read_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip().lower().replace(" ", "_") for cell in ws[1]]
            if "ip_address" not in headers:
                return jsonify({"error": "XLSX must have an 'ip_address' column"}), 400

            imported = []
            skipped = []
            rows_data = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i]:
                        row_dict[headers[i]] = str(val).strip() if val else ""
                ip = row_dict.get("ip_address", "").strip()
                if not ip:
                    continue
                if db.switches.find_one({"ip_address": ip}):
                    skipped.append({"ip_address": ip, "reason": "IP already exists"})
                    continue

                doc = switch_schema({
                    "name": row_dict.get("name") or row_dict.get("hostname") or f"SW-{ip.replace('.', '-')}",
                    "ip_address": ip,
                    "model": row_dict.get("model", ""),
                    "platform": row_dict.get("platform", "IOS-XE"),
                    "current_version": row_dict.get("version") or row_dict.get("current_version", ""),
                    "serial_number": row_dict.get("serial_number") or row_dict.get("serial", ""),
                    "site": row_dict.get("site") or row_dict.get("location", ""),
                    "ssh_username": row_dict.get("ssh_username") or row_dict.get("username", "admin"),
                    "ssh_password": row_dict.get("ssh_password") or row_dict.get("password", ""),
                    "enable_password": row_dict.get("enable_password") or row_dict.get("enable", ""),
                    "status": "discovering",
                    "notes": row_dict.get("notes", ""),
                })
                result = db.switches.insert_one(doc)
                doc["_id"] = str(result.inserted_id)
                imported.append(doc)
                rows_data.append({
                    "id": str(result.inserted_id), "ip": ip,
                    "username": doc["ssh_username"], "password": doc["ssh_password"],
                    "enable": doc["enable_password"],
                })
            wb.close()
        except Exception as e:
            return jsonify({"error": f"Failed to parse XLSX: {str(e)}"}), 400

        # Create discovery job tracker
        discovery_id = str(uuid.uuid4())
        with _discovery_lock:
            _discovery_jobs[discovery_id] = {
                "id": discovery_id,
                "total": len(rows_data),
                "completed": 0,
                "succeeded": 0,
                "failed": 0,
                "status": "running",
                "switches": {sd["ip"]: {"ip": sd["ip"], "id": sd["id"], "status": "pending", "detail": ""} for sd in rows_data},
            }

        if rows_data:
            def _auto_discover_all():
                import logging
                log = logging.getLogger("bulk_import")
                from concurrent.futures import ThreadPoolExecutor, as_completed

                def _discover_one(sw_data):
                    sw_id = sw_data["id"]
                    ip = sw_data["ip"]
                    update = {"updated_at": datetime.now(timezone.utc)}

                    # Update tracker: discovering
                    with _discovery_lock:
                        _discovery_jobs[discovery_id]["switches"][ip]["status"] = "discovering"
                        _discovery_jobs[discovery_id]["switches"][ip]["detail"] = "Connecting via SSH…"

                    try:
                        from netmiko import ConnectHandler
                        conn = ConnectHandler(
                            device_type="cisco_xe", host=ip,
                            username=sw_data["username"], password=sw_data["password"],
                            secret=sw_data["enable"], timeout=15, conn_timeout=10,
                        )
                        conn.enable()

                        with _discovery_lock:
                            _discovery_jobs[discovery_id]["switches"][ip]["detail"] = "Reading hostname…"

                        output = conn.send_command("show run | include ^hostname")
                        if output.strip():
                            update["name"] = output.strip().replace("hostname ", "").strip()

                        with _discovery_lock:
                            _discovery_jobs[discovery_id]["switches"][ip]["detail"] = "Reading show version…"

                        ver = conn.send_command("show version")
                        conn.disconnect()

                        for line in ver.splitlines():
                            ll = line.lower()
                            if "model number" in ll or ("cisco " in ll and "processor" in ll):
                                for p in line.split():
                                    if p.startswith(("C9", "C38", "C36", "C35", "N9K", "WS-")):
                                        update["model"] = p
                            if "version" in ll and ("ios" in ll or "nx-os" in ll):
                                for p in line.split():
                                    if p[0:1].isdigit():
                                        update["current_version"] = p.strip(",").strip()
                                        break
                                if "nx-os" in ll:
                                    update["platform"] = "NX-OS"
                                else:
                                    update["platform"] = "IOS-XE"
                            if "board id" in ll or "system serial" in ll:
                                update["serial_number"] = line.split()[-1]

                        update["status"] = "online"
                        update["last_seen"] = datetime.now(timezone.utc)
                        db.switches.update_one({"_id": ObjectId(sw_id)}, {"$set": update})

                        detail = f"{update.get('name', ip)} — {update.get('model', '?')} — v{update.get('current_version', '?')}"
                        log.info(f"Discovered {ip}: {detail}")

                        with _discovery_lock:
                            job = _discovery_jobs[discovery_id]
                            job["completed"] += 1
                            job["succeeded"] += 1
                            job["switches"][ip]["status"] = "success"
                            job["switches"][ip]["detail"] = detail

                    except Exception as e:
                        update["status"] = "offline"
                        update["notes"] = f"Auto-discover failed: {str(e)[:120]}"
                        db.switches.update_one({"_id": ObjectId(sw_id)}, {"$set": update})
                        log.warning(f"Discover failed for {ip}: {e}")

                        with _discovery_lock:
                            job = _discovery_jobs[discovery_id]
                            job["completed"] += 1
                            job["failed"] += 1
                            job["switches"][ip]["status"] = "failed"
                            job["switches"][ip]["detail"] = str(e)[:100]

                with ThreadPoolExecutor(max_workers=10) as pool:
                    futures = [pool.submit(_discover_one, sd) for sd in rows_data]
                    for f in as_completed(futures):
                        try:
                            f.result()
                        except Exception as e:
                            log.error(f"Discovery thread error: {e}")

                with _discovery_lock:
                    _discovery_jobs[discovery_id]["status"] = "complete"
                log.info(f"Bulk discovery complete: {len(rows_data)} switches")

            _threading.Thread(target=_auto_discover_all, daemon=True).start()

        return jsonify({
            "imported": len(imported),
            "skipped": len(skipped),
            "skipped_details": skipped,
            "discovery_id": discovery_id,
            "message": f"Imported {len(imported)} switches. Discovery started.",
        }), 201

    @switches_bp.route("/discovery-progress/<disc_id>", methods=["GET"])
    def discovery_progress(disc_id):
        """Poll discovery progress."""
        with _discovery_lock:
            job = _discovery_jobs.get(disc_id)
        if not job:
            return jsonify({"error": "Not found"}), 404
        return jsonify(job)

    @switches_bp.route("/discovery-progress/<disc_id>/stream", methods=["GET"])
    def discovery_stream(disc_id):
        """SSE stream for discovery progress."""
        import json, time
        from flask import Response
        def gen():
            while True:
                with _discovery_lock:
                    job = _discovery_jobs.get(disc_id)
                if not job:
                    yield f"data: {json.dumps({'status': 'not_found'})}\n\n"
                    break
                yield f"data: {json.dumps(job, default=str)}\n\n"
                if job["status"] == "complete":
                    break
                time.sleep(2)
        return Response(gen(), mimetype="text/event-stream",
                        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

    @switches_bp.route("/template", methods=["GET"])
    def download_template():
        """Download an XLSX template for bulk import."""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from flask import send_file

        wb = Workbook()
        ws = wb.active
        ws.title = "Switches"

        # Headers
        headers = ["ip_address", "ssh_username", "ssh_password", "enable_password", "site", "notes"]
        header_fill = PatternFill(start_color="C8462B", end_color="C8462B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Example rows
        examples = [
            ["10.0.1.1", "admin", "cisco123", "enable123", "HQ-DataCenter", "Core switch"],
            ["10.0.1.2", "admin", "cisco123", "enable123", "HQ-Floor2", ""],
            ["10.0.2.10", "admin", "cisco123", "", "Branch-Office-1", "Access switch"],
        ]
        for row_idx, ex in enumerate(examples, 2):
            for col, val in enumerate(ex, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border

        # Column widths
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 25

        # Add instructions sheet
        ws2 = wb.create_sheet("Instructions")
        instructions = [
            "Cisco Switch Bulk Import Template",
            "",
            "Required columns:",
            "  ip_address — Switch management IP (required)",
            "  ssh_username — SSH login username",
            "  ssh_password — SSH login password",
            "",
            "Optional columns:",
            "  enable_password — Enable/secret password",
            "  site — Site/location name",
            "  notes — Any notes",
            "",
            "Optional (auto-detected via SSH if left blank):",
            "  name / hostname",
            "  model",
            "  platform (IOS-XE / NX-OS / IOS)",
            "  version / current_version",
            "  serial_number",
            "",
            "After import, the app will SSH to each switch",
            "and auto-detect: hostname, model, version, serial, platform.",
        ]
        for i, line in enumerate(instructions, 1):
            cell = ws2.cell(row=i, column=1, value=line)
            if i == 1:
                cell.font = Font(bold=True, size=14, color="C8462B")
            elif line.startswith("  "):
                cell.font = Font(name="Consolas", size=10)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="switch_import_template.xlsx")

    @switches_bp.route("/discover", methods=["POST"])
    def discover():
        d = request.json
        ip = d.get("ip_address", "")
        username = d.get("ssh_username", "admin")
        password = d.get("ssh_password", "")
        if not ip: return jsonify({"error": "ip_address required"}), 400

        hostname = ""

        # 1) Try SSH to get the real configured hostname
        try:
            from netmiko import ConnectHandler
            conn = ConnectHandler(
                device_type="cisco_xe", host=ip,
                username=username, password=password,
                secret=d.get("enable_password", ""),
                timeout=10
            )
            conn.enable()

            # Get hostname from running config
            output = conn.send_command("show run | include ^hostname")
            if output.strip():
                hostname = output.strip().replace("hostname ", "").strip()

            # Also grab model, version, serial from 'show version'
            ver_output = conn.send_command("show version")
            conn.disconnect()

            model = ""
            version = ""
            serial = ""
            platform = "IOS-XE"
            for line in ver_output.splitlines():
                ll = line.lower()
                if "model number" in ll or "cisco " in ll and "processor" in ll:
                    parts = line.split()
                    for p in parts:
                        if p.startswith("C9") or p.startswith("C38") or p.startswith("C36") or p.startswith("N9K"):
                            model = p
                if "version" in ll and ("ios" in ll or "nx-os" in ll):
                    for p in line.split():
                        if p[0:1].isdigit():
                            version = p.strip(",").strip()
                            break
                    if "nx-os" in ll:
                        platform = "NX-OS"
                if "board id" in ll or "system serial" in ll:
                    serial = line.split()[-1]

            return jsonify({
                "name": hostname or f"SW-{ip.replace('.', '-')}",
                "ip_address": ip,
                "model": model or "Unknown",
                "platform": platform,
                "current_version": version or "Unknown",
                "serial_number": serial or "",
                "ssh_username": username,
                "status": "online",
            })

        except Exception as ssh_err:
            # 2) SSH failed — fallback to reverse DNS
            import socket
            if not hostname:
                try:
                    hostname = socket.gethostbyaddr(ip)[0]
                except (socket.herror, socket.gaierror, OSError):
                    hostname = f"SW-{ip.replace('.', '-')}"

            return jsonify({
                "name": hostname,
                "ip_address": ip,
                "model": "Unknown (SSH failed)",
                "platform": "IOS-XE",
                "current_version": "Unknown",
                "serial_number": "",
                "ssh_username": username,
                "status": "unknown",
                "notes": f"Auto-discover SSH failed: {str(ssh_err)[:120]}",
            })

    # ── Dynamic /<sid> paths AFTER all static paths ──────────

    @switches_bp.route("/<sid>", methods=["GET"])
    def get_one(sid):
        sw = db.switches.find_one({"_id": ObjectId(sid)})
        if not sw: return jsonify({"error": "Not found"}), 404
        sw["_id"] = str(sw["_id"])
        return jsonify(sw)

    @switches_bp.route("/<sid>", methods=["PUT"])
    def update(sid):
        d = request.json
        fields = {}
        for k in ["name","ip_address","model","platform","current_version",
                   "serial_number","site","ssh_username","ssh_password",
                   "enable_password","status","notes",
                   "is_stack","stack_count","stack_master","stack_members"]:
            if k in d: fields[k] = d[k]
        fields["updated_at"] = datetime.now(timezone.utc)
        db.switches.update_one({"_id": ObjectId(sid)}, {"$set": fields})
        sw = db.switches.find_one({"_id": ObjectId(sid)})
        sw["_id"] = str(sw["_id"])
        return jsonify(sw)

    @switches_bp.route("/<sid>", methods=["DELETE"])
    def delete(sid):
        r = db.switches.delete_one({"_id": ObjectId(sid)})
        if r.deleted_count == 0: return jsonify({"error": "Not found"}), 404
        return jsonify({"message": "Deleted"})

    @switches_bp.route("/<sid>/check", methods=["POST"])
    def check_one(sid):
        """Ping a single switch and update its status."""
        from services.health_checker import check_switch
        sw = db.switches.find_one({"_id": ObjectId(sid)})
        if not sw:
            return jsonify({"error": "Switch not found"}), 404
        timeout = int(request.args.get("timeout", 2))
        result = check_switch(db, sw, ping_timeout=timeout)
        return jsonify(result)

    return switches_bp